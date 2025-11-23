import json
import openpyxl
import os
import logging
import shutil
from commonomuhelper.omuhelper import produce_output_parameters
from datetime import datetime
from openpyxl.reader.excel import load_workbook
from os.path import join, getctime
from typing import List

from bobbtransfojob.common.common_constants import EXCEL_FILE_EXTENSION, JSON_FILE_EXTENSION, LIST_SELECTED_FILES, \
    EMPTY_STRING, LIST_INPUT_FILES, NUMBER_INPUT_FILES, \
    LIST_ORDERED_FILES, NUMBER_SELECTED_FILES, USER_ID, CODE_CONTRACT_ELEMENT, CODE_INSURER, PRODUCT_ELEMENTS, IGNORED, \
    CODE_AMC, CODE_OC, CODE_OFFER, CODE_PRODUCT, CODE_BENEFIT_NATURE, FROM_DATE, EFFECTIVE_DATE, TO_DATE, CHUNK_SIZE, \
    ASC, DESC
from bobbtransfojob.models.contract_element import ContractElement
from bobbtransfojob.models.excel_column import COLUMN_CODE_OC, COLUMN_CODE_OFFER, COLUMN_CODE_PRODUCT, \
    COLUMN_CODE_BENEFIT_NATURE, COLUMN_FROM_DATE, COLUMN_TO_DATE, COLUMN_EFFECTIVE_DATE, COLUMN_CODE_CONTRACT_ELEMENT, \
    COLUMN_CODE_INSURER, COLUMN_CODE_AMC, COLUMN_IGNORED
from bobbtransfojob.models.product_element import ProductElement
from bobbtransfojob.services.event_publisher_service import EventPublisherService
from bobbtransfojob.exceptions.excel_not_sorted_exception import ExcelNotSortedException
from bobbtransfojob.logging import init_logging

LOGGER_NAME = "bobbtransfo"
init_logging(LOGGER_NAME)
logger = logging.getLogger(LOGGER_NAME)


class ExcelToJsonConverterService:
    def __init__(self, input_folder: str, output_folder: str):
        self.input_folder = input_folder
        self.output_folder = output_folder
        self.list_selected_files = []
        self.list_untransformed_file = []
        self.list_json_file = []
        self.list_input_files = []
        self.list_ordered_files = []
        self.user_id = os.getenv(USER_ID)
        self.batch_size = int(os.getenv(CHUNK_SIZE, "1000"))

    @staticmethod
    def delete_file(file_path):
        try:
            os.remove(file_path)
            logger.info(f"Deleted file: {file_path}")
        except Exception as del_error:
            logger.error(f"Error occurred while deleting file: {file_path}. Error: {str(del_error)}")

    @staticmethod
    def is_excel_file_empty(file_path):
        try:
            workbook = load_workbook(filename=file_path, read_only=True)
            sheet = workbook.active
            rows = sheet.iter_rows(min_row=1, max_row=2, values_only=True)
            num_rows = sum(1 for _ in rows)
            return num_rows < 2
        except Exception as e:
            logger.error(f"Error occurred while checking Excel file: {e}")
            return True

    @staticmethod
    def is_json_file_empty(file_path):
        with open(file_path, 'r') as file:
            # Read the file in small chunks
            for chunk in iter(lambda: file.read(1024), ''):
                # Remove any surrounding whitespace from the chunk
                stripped_chunk = chunk.strip()
                if stripped_chunk:  # If any non-whitespace characters are present
                    return False  # There's some JSON content
        return True  # File is considered empty if no JSON content found

    @staticmethod
    def handle_publish_failed_event(error_msg):
        event_publisher = EventPublisherService()
        event_publisher.publish_failed_event(error_msg)

    def handle_empty_file(self, filename, file_path):
        msg = f"The file {filename} is empty. Skipping processing."
        logger.error(msg)
        self.handle_publish_failed_event(msg)
        self.delete_file(file_path)

    def get_files_to_process(self, input_file):
        files = []
        if self.user_id:
            files.append(input_file)
        else:
            files = sorted(os.listdir(self.input_folder), key=lambda x: getctime(join(self.input_folder, x)))
        return files

    def process_files_in_folder(self, input_file):
        try:
            if not os.path.exists(self.output_folder):
                os.makedirs(self.output_folder)

            files = self.get_files_to_process(input_file)

            for filename in files:
                logger.info(f"=== Begin process file {filename} :")
                file_path = os.path.join(self.input_folder, filename)
                if filename.endswith(EXCEL_FILE_EXTENSION):
                    self.list_input_files.append(file_path)
                    if self.is_excel_file_empty(file_path):
                        self.handle_empty_file(filename, file_path)
                        self.list_untransformed_file.append(file_path)
                    else:
                        self.process_excel_file(filename)
                elif filename.endswith(JSON_FILE_EXTENSION):
                    self.list_input_files.append(file_path)
                    if self.is_json_file_empty(file_path):
                        self.handle_empty_file(filename, file_path)
                    else:
                        self.process_json_file(filename)
                else:
                    logger.warning(f"Ignored file: {filename}")

        except ExcelNotSortedException as e:
            raise e
        except Exception as e:
            logger.error(f"An error occurred: {str(e)}", exc_info=True)
            self.handle_publish_failed_event(str(e))

    def process_excel_file(self, filename):
        try:
            excel_file_path = os.path.join(self.input_folder, filename)
            if self.check_excel_columns(excel_file_path):
                self.excel_to_json(excel_file_path, filename)
                output_filename = os.path.splitext(filename)[0] + JSON_FILE_EXTENSION
                logger.info(f"Converted EXCEL: {filename} to JSON: {output_filename}")
                event_publisher = EventPublisherService()
                event_publisher.publish_success_event(filename)
                self.list_selected_files.append(filename)
                self.list_ordered_files.append(output_filename)
            else:
                self.list_untransformed_file.append(excel_file_path)

        except ExcelNotSortedException as e:
            self.handle_error(e)
            self.handle_publish_failed_event(str(e))
            raise e
        except Exception as e:
            msg = f"An error occurred while processing Excel file: {str(e)}"
            logger.error(msg, exc_info=True)
            self.handle_publish_failed_event(msg)

    def process_json_file(self, filename):
        try:
            json_file_path = os.path.join(self.input_folder, filename)
            output_file_path = os.path.join(self.output_folder, filename)
            shutil.copy(json_file_path, output_file_path)
            logger.info(f"Copied JSON file to output folder: {output_file_path}")
            self.list_json_file.append(filename)
            self.list_ordered_files.append(filename)

        except Exception as e:
            msg = f"An error occurred while processing JSON file: {str(e)}"
            logger.error(msg, exc_info=True)

    def excel_to_json(self, input_file_path, filename):
        try:
            output_filename = os.path.splitext(filename)[0] + ".json"
            output_file_path = os.path.join(self.output_folder, output_filename)
            logger.info("Start parsing to json")

            with open(output_file_path, 'w') as json_file:
                json_file.write('[')
                first_group = True

                last_contract_element = None
                last_insurer = None
                current_group = []
                order = None

                for batch_num, batch in enumerate(self.read_excel_file_in_batches(input_file_path, self.batch_size)):
                    logger.info(f"Parsing batch {batch_num + 1}...")
                    for row in batch:
                        current_contract_element = row[CODE_CONTRACT_ELEMENT]
                        current_insurer = row[CODE_INSURER]

                        is_ordered, order = self.determine_and_validate_order(
                            last_contract_element, last_insurer,
                            current_contract_element, current_insurer,
                            order
                        )
                        if not is_ordered:
                            msg = f"Le fichier {filename} doit être trié par codecontract element+ code insurer."
                            self.list_untransformed_file.append(input_file_path)
                            raise ExcelNotSortedException(msg)

                        if (last_contract_element is not None and last_insurer is not None and
                                (current_contract_element != last_contract_element or current_insurer != last_insurer)):
                            json_data = self.process_group(current_group)
                            if not first_group:
                                json_file.write(',\n')
                            json.dump(json_data, json_file, indent=4)
                            first_group = False
                            current_group = []

                        current_group.append(row)
                        last_contract_element = current_contract_element
                        last_insurer = current_insurer

                if current_group:
                    json_data = self.process_group(current_group)
                    if not first_group:
                        json_file.write(',\n')
                    json.dump(json_data, json_file, indent=4)

                json_file.write(']')
                logger.info("End parsing to json")
        except Exception as e:
            msg = f"An error occurred while converting CSV to JSON: {str(e)}"
            logger.error(msg, exc_info=True)
            self.handle_publish_failed_event(msg)

    def read_excel_file_in_batches(self, file_path, batch_size=1000):
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet = workbook.active
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        batch = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(row):
                batch.append(dict(zip(header, row)))
                if len(batch) == batch_size:
                    yield batch
                    batch = []
        if batch:
            yield batch

    def process_group(self, group):
        def parse_date(date_str):
            if not date_str:
                return EMPTY_STRING
            if isinstance(date_str, datetime):
                return date_str.strftime("%Y-%m-%d")
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y"):
                try:
                    return datetime.strptime(date_str, fmt).strftime("%Y-%m-%d")
                except ValueError:
                    continue
            return date_str

        def create_product_element(item):
            return {
                CODE_OC: str(item[CODE_OC]) if item[CODE_OC] else EMPTY_STRING,
                CODE_OFFER: str(item[CODE_OFFER]) if item[CODE_OFFER] else EMPTY_STRING,
                CODE_PRODUCT: str(item[CODE_PRODUCT]) if item[CODE_PRODUCT] else EMPTY_STRING,
                CODE_BENEFIT_NATURE: str(item[CODE_BENEFIT_NATURE]) if item[CODE_BENEFIT_NATURE] else EMPTY_STRING,
                FROM_DATE: parse_date(item[FROM_DATE]) if item[FROM_DATE] else EMPTY_STRING,
                EFFECTIVE_DATE: parse_date(item[EFFECTIVE_DATE]) if item[EFFECTIVE_DATE] else EMPTY_STRING,
                TO_DATE: parse_date(item[TO_DATE]) if item[TO_DATE] else EMPTY_STRING
            }

        def is_required_column_present(item):
            required_columns = [
                CODE_OC,
                CODE_OFFER,
                CODE_PRODUCT,
                CODE_BENEFIT_NATURE,
                FROM_DATE,
                TO_DATE,
                EFFECTIVE_DATE
            ]
            return not all(not item[column] for column in required_columns)

        contract_element = group[0][CODE_CONTRACT_ELEMENT]
        insurer = group[0][CODE_INSURER]
        amc = group[0][CODE_AMC]
        ignored = group[0][IGNORED]
        product_elements = [create_product_element(item) for item in group if is_required_column_present(item)]

        return {
            CODE_CONTRACT_ELEMENT: str(contract_element) if contract_element else EMPTY_STRING,
            CODE_INSURER: str(insurer) if insurer else EMPTY_STRING,
            CODE_AMC: str(amc) if amc else EMPTY_STRING,
            PRODUCT_ELEMENTS: product_elements,
            IGNORED: str(ignored) if ignored else EMPTY_STRING
        }

    def check_excel_columns(self, excel_file_path):
        try:
            wb = openpyxl.load_workbook(excel_file_path, read_only=True)
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            required_columns = [
                COLUMN_CODE_CONTRACT_ELEMENT.name, COLUMN_CODE_INSURER.name, COLUMN_CODE_AMC.name,
                COLUMN_CODE_OC.name, COLUMN_CODE_OFFER.name, COLUMN_CODE_PRODUCT.name,
                COLUMN_CODE_BENEFIT_NATURE.name, COLUMN_FROM_DATE.name, COLUMN_TO_DATE.name,
                COLUMN_EFFECTIVE_DATE.name, COLUMN_IGNORED.name
            ]
            missing_columns = [column for column in required_columns if column not in header_row]
            if missing_columns:
                msg = f"Missing columns in Excel file: {missing_columns}"
                logger.error(msg)
                self.handle_publish_failed_event(msg)
                return False
            return True
        except Exception as e:
            logger.error(f"An error occurred while checking Excel columns: {str(e)}", exc_info=True)
            return False

    def create_product_element(self, row) -> ProductElement:
        try:
            required_columns = [
                COLUMN_CODE_OC.index,
                COLUMN_CODE_OFFER.index,
                COLUMN_CODE_PRODUCT.index,
                COLUMN_CODE_BENEFIT_NATURE.index,
                COLUMN_FROM_DATE.index,
                COLUMN_TO_DATE.index,
                COLUMN_EFFECTIVE_DATE.index
            ]

            if all(not row[column] for column in required_columns):
                return None

            product_element = ProductElement(
                code_oc=row[COLUMN_CODE_OC.index] or EMPTY_STRING,
                code_offer=row[COLUMN_CODE_OFFER.index] or EMPTY_STRING,
                code_product=row[COLUMN_CODE_PRODUCT.index] or EMPTY_STRING,
                code_benefit_nature=row[COLUMN_CODE_BENEFIT_NATURE.index] or EMPTY_STRING,
                from_date=row[COLUMN_FROM_DATE.index] or EMPTY_STRING,
                to_date=row[COLUMN_TO_DATE.index] or EMPTY_STRING,
                effective_date=row[COLUMN_EFFECTIVE_DATE.index] or EMPTY_STRING
            )
            return product_element
        except Exception:
            return None

    def create_contract_element(self, row, product_elements: List[ProductElement]) -> ContractElement:
        try:
            contract_element = ContractElement(
                code_contract_element=row[COLUMN_CODE_CONTRACT_ELEMENT.index] if row[
                    COLUMN_CODE_CONTRACT_ELEMENT.index] else EMPTY_STRING,
                code_insurer=row[COLUMN_CODE_INSURER.index] if row[COLUMN_CODE_INSURER.index] else EMPTY_STRING,
                code_amc=row[COLUMN_CODE_AMC.index] if row[COLUMN_CODE_AMC.index] else EMPTY_STRING,
                product_elements=product_elements,
                ignored=row[COLUMN_IGNORED.index] if row[COLUMN_IGNORED.index] else EMPTY_STRING
            )
            return contract_element
        except Exception:
            return None

    def generate_crex(self):
        produce_output_parameters({
            LIST_INPUT_FILES: self.list_input_files,
            NUMBER_INPUT_FILES: len(self.list_input_files),
            LIST_ORDERED_FILES: self.list_ordered_files,
            LIST_SELECTED_FILES: self.list_selected_files + self.list_json_file,
            NUMBER_SELECTED_FILES: len(self.list_selected_files + self.list_json_file),
            USER_ID: os.getenv(USER_ID, "")
        })
        logger.info(f"list_input_files : {self.list_input_files}")
        logger.info(f"number_input_files : {len(self.list_input_files)}")
        logger.info(f"list_selected_files : {self.list_selected_files + self.list_json_file}")
        logger.info(f"number_selected_files : {len(self.list_selected_files + self.list_json_file)}")

    def is_excel_sorted(self, file_path, order, batch_size=1000):
        last_contract_element = None
        last_insurer = None

        for batch in self.read_excel_file_in_batches(file_path, batch_size):
            for row in batch:
                current_contract_element = row[CODE_CONTRACT_ELEMENT]
                current_insurer = row[CODE_INSURER]

                if self.is_out_of_order(last_contract_element, last_insurer, current_contract_element, current_insurer,
                                        order):
                    logger.debug(f"order : {order}")
                    logger.debug(f"last_contract_element : {last_contract_element}")
                    logger.debug(f"last_insurer : {last_insurer}")
                    logger.debug(f"current_contract_element : {current_contract_element}")
                    logger.debug(f"current_insurer : {current_insurer}")
                    return False
                last_contract_element, last_insurer = current_contract_element, current_insurer

        return True

    def load_sheet(self, file_path):
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        return workbook.active

    def get_column_indices(self, sheet):
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        return header.index(CODE_CONTRACT_ELEMENT), header.index(CODE_INSURER)

    def is_out_of_order(self, last_contract_element, last_insurer, current_contract_element, current_insurer, order):
        if last_contract_element is None or last_insurer is None:
            return False  # No comparison needed for the first row

        current_contract_element = self.normalize(current_contract_element)
        last_contract_element = self.normalize(last_contract_element)
        current_insurer = self.normalize(current_insurer)
        last_insurer = self.normalize(last_insurer)

        if type(current_contract_element) != type(last_contract_element):
            if isinstance(current_contract_element, (int, float)):
                return False if order == ASC else True
            else:
                return True if order == ASC else False

        if order == ASC:
            return (current_contract_element < last_contract_element or
                    (current_contract_element == last_contract_element and current_insurer < last_insurer))
        elif order == DESC:
            return (current_contract_element > last_contract_element or
                    (current_contract_element == last_contract_element and current_insurer > last_insurer))

        return False

    def normalize(self, value):
        if value is None:
            return ""
        value = str(value).strip().lower()

        if value.isdigit():
            return int(value)
        try:
            return float(value)
        except ValueError:
            return value



    def handle_error(self, exception):
        logger.error(f"An error occurred while checking sorting order: {str(exception)}", exc_info=True)

    def determine_and_validate_order(self,
                                     last_contract_element,
                                     last_insurer,
                                     current_contract_element,
                                     current_insurer,
                                     order):
        if order is None and last_contract_element is not None and last_insurer is not None:
            if current_contract_element > last_contract_element or (
                    current_contract_element == last_contract_element and current_insurer > last_insurer):
                order = ASC
            elif current_contract_element < last_contract_element or (
                    current_contract_element == last_contract_element and current_insurer < last_insurer):
                order = DESC

        if last_contract_element is not None and last_insurer is not None:
            if self.is_out_of_order(last_contract_element, last_insurer, current_contract_element, current_insurer, order):
                logger.debug("Sorting order failed during processing")
                return False, order

        return True, order

